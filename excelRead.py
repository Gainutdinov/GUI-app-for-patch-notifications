from openpyxl import load_workbook
import os
import datetime

def readExcelFile(path, patchDate):
    #print(os.getcwd())
    serversToPatch={}
    wb = load_workbook(filename = path)
    sheet_ranges = wb[wb.active.title]
    shtLstRow=sheet_ranges.max_row
    # print('Number of rows found - ',shtLstRow)
    for _ in sheet_ranges[1]:
        if 'Next maintenance window start' in _.value:
            NxtMnt=sheet_ranges[1].index(_)
            print(_.value)
            break
    for counter,value in enumerate(sheet_ranges):
        cell=value[NxtMnt].value
        if type(cell) is datetime.datetime:
            dat=cell.date().strftime('%m/%d/%Y')
            if dat==patchDate:
                # print('Server ->',value[0].value,'-- Next maintenance window start->',value[NxtMnt].value,'--Next maintenance window end->',value[NxtMnt+1].value)
                serversToPatch[value[0].value]=(str(value[NxtMnt].value),str(value[NxtMnt+1].value))
    return serversToPatch